import openpyxl
from datetime import datetime, timedelta
import json


def convert_excel_to_json(sheet):
    jsondata = []

    try:

        # 遍历每一列
        coldatas = [list(column) for column in sheet.iter_cols(values_only=True)]
        sheetname = sheet.title

        # 按下标遍历
        datetimevalue = coldatas[1][1]
        coldatas[1][1] = datetimevalue.strftime("%Y/%m/%d")
        # print('datetimevalue:', datetimevalue)
        n = 0
        for i in range(3, len(coldatas), 2):
            # value = coldatas[i][1]
            n = n + 1
            datetimevalue_n = timedelta(days=n) + datetimevalue
            datetimevalue_n_str = datetimevalue_n.strftime("%Y/%m/%d")
            coldatas[i][1] = datetimevalue_n_str
            # print(coldatas[i][1])
            # datetimevalue = datetime.strptime(datetimevalue, "%Y-%m-%d %H:%M:%S")
            # print(value)
            # prevalue = coldatas[i-2][1]
            # if isinstance(value, str) and '=' in value:
            #     # value = datetime.strptime(prevalue, "%Y-%m-%d %H:%M:%S") 
            #     # print(value)
            #     pass
        # print(coldatas)
        # del coldatas[0][0]
        for i in range(1, len(coldatas[0]), 1):
            roomdata = {}
            # print(coldatas[0][i])
            value = coldatas[0][i]
            if not value is None:
                roomdata['room'] = value
                roomdata['cells'] = []
                for j in range(1, len(coldatas),2):
                    cell = {}
                    date = coldatas[j][1]
                    name = coldatas[j][i]
                    if name is None:
                        continue
                    price = coldatas[j+1][i]
                    cell['date'] = date
                    cell['name'] = name
                    cell['price'] = price
                    #如果price不可以转为数字,则跳过
                    try:
                        cell['price'] = float(cell['price'])
                    except:
                        continue
                    roomdata['cells'].append(cell)
                jsondata.append(roomdata)
        # print(coldatas)

        #保存fdata到json
        json_str = json.dumps(jsondata, ensure_ascii=False)
        with open(f'{sheetname}.json', 'w', encoding='utf-8') as f:
            f.write(json_str)
        print(f'{sheetname}.json文件已生成')

    except FileNotFoundError:
        print(f"错误: 文件 {file_path} 未找到。")
    except Exception as e:
        print(f"错误: 发生了未知错误: {e}")


def read_excel_column_by_column(file_path):

    # 加载工作簿
    workbook = openpyxl.load_workbook(file_path)
    # 获取所有工作表
    sheets = workbook.sheetnames
    for sheet in sheets:
        sheet = workbook[sheet]
        convert_excel_to_json(sheet)


if __name__ == "__main__":
    file_path = "/Users/fgmac/Downloads/入住登记表/入住登记表月份分离.xlsx"
    read_excel_column_by_column(file_path)